import openpyxl
# from ..models import *
from ..models.daily_report_history import DailyReport
from django.contrib.auth import get_user_model
from .whatsapp_services import assign_telecaller
from ..models.leads import Lead
from rest_framework.exceptions import APIException
from .query_services import exec_raw_sql
# from ..views.dynamic_pdf import *
from django.utils import timezone
from datetime import datetime
from ..services.dashboard_services import get_user_details
from ..tasks.notification_task import send_lead_assigned_notification



# def lead_upload_excel(**data):
#     file_obj = data.get('file')

#     wb = openpyxl.load_workbook(file_obj)
#     sheet = wb.active

#     already_assign = []
#     success_count = 0      # Newly inserted count
#     duplicate_count = 0    # Already exists count

#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if any(row):

#             telecaller = assign_telecaller()

#             full_name, mobile_no, lead_source = row[:3]

#             lead = Lead.objects.filter(mobile_no=mobile_no).first()

#             if lead:
#                 duplicate_count += 1

#                 already_assign.append(
#                     f"{mobile_no} already exists. Assigned to {lead.assigned_to}"
#                 )

#             else:
#                 Lead.objects.create(
#                     full_name=full_name,
#                     mobile_no=mobile_no,
#                     current_status="Working",
#                     pipeline_stage_id=1,
#                     priority_id=4,
#                     assigned_to=telecaller
#                 )

#                 success_count += 1

#     return {
#         "total_rows": sheet.max_row - 1,
#         "success_count": success_count,
#         "duplicate_count": duplicate_count,
#         "duplicates": already_assign,
#         "message": f"{success_count} leads inserted successfully."
#     }
    






def lead_upload_excel(**data):
    """
    Same save logic as lead_upload_excel(), but takes an already-parsed
    JSON list of leads (from the frontend preview/edit step) instead of
    reading an excel file.
 
    Expected `data`:
    {
        "leads": [
            {"full_name": "Ravi Kumar", "mobile_no": "9876543213", "lead_source": "Insta"},
            ...
        ]
    }
    """
 
    leads = data.get("leads") or []
    print(leads)
 
    already_assign = []
    success_count = 0      # Newly inserted count
    duplicate_count = 0    # Already exists count (re-checked here, at save time)
 
    for item in leads:
 
        full_name = item.get("full_name")
        mobile_no = item.get("mobile_no")
        lead_source_id = item.get("lead_source")
        campaign = item.get("campaign_name")
 
        # re-check duplicate at save time too — DB could have changed
        # since the preview step ran
        lead = Lead.objects.filter(mobile_no=mobile_no).first()
 
        if lead:
            duplicate_count += 1
 
            already_assign.append(
                f"{mobile_no} already exists. Assigned to {lead.assigned_to}"
            )
            continue
 
        telecaller = assign_telecaller()
 
        lead=Lead.objects.create(
            full_name=full_name,
            mobile_no=mobile_no,
            current_status="Working",
            pipeline_stage_id=1,
            priority_id=4,
            assigned_to=telecaller,
            lead_source_id=lead_source_id,
            campaign_id=campaign
        )
        send_lead_assigned_notification(lead.id, lead.assigned_to,assigned_by_id=None)
        success_count += 1
 
    return {
        "total_rows": len(leads),
        "success_count": success_count,
        "duplicate_count": duplicate_count,
        "duplicates": already_assign,
        "message": f"{success_count} leads inserted successfully.",
    }


def preview_lead_excel(**data):

    file_obj = data.get("file")

    wb = openpyxl.load_workbook(file_obj)
    sheet = wb.active

    preview_data = []

    total_rows = 0
    valid_rows = 0
    duplicate_rows = 0
    error_rows = 0

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

        if not any(row):
            continue

        total_rows += 1

        full_name = row[0]
        mobile_no = str(row[1]).strip() if row[1] else ""
        lead_source = row[2]
        campaign_name=row[3]

        status = "Valid"
        message = ""

        if not full_name:
            status = "Error"
            message = "Full Name Required"

        elif not mobile_no:
            status = "Error"
            message = "Mobile Number Required"

        elif len(mobile_no) != 10:
            status = "Error"
            message = "Invalid Mobile Number"

        else:

            lead = Lead.objects.filter(mobile_no=mobile_no).first()

            if lead:

                status = "Duplicate"
                message = f"Already Assigned to {lead.assigned_to}"

        if status == "Valid":
            valid_rows += 1

        elif status == "Duplicate":
            duplicate_rows += 1

        else:
            error_rows += 1

        preview_data.append({

            "row": index,

            "full_name": full_name,

            "mobile_no": mobile_no,

            "lead_source": lead_source,
            
            "campaign_name":campaign_name,

            "status": status,

            "message": message

        })

    return {

        "total_rows": total_rows,

        "valid_rows": valid_rows,

        "duplicate_rows": duplicate_rows,

        "error_rows": error_rows,

        "preview_data": preview_data

    }






def daily_report(**data):
    try:
        user_id = data.get("id")

        User = get_user_model() 

        tele = User.objects.filter(id=user_id).first()
        if tele is None:
            raise APIException("Telecaller Not Found")

        # 🔥 correct param passing
        report = exec_raw_sql(
            "D_FETCH_DAILY_REPORT_PDF",
            {"id": user_id}
        )
        
        latest_report = DailyReport.objects.filter(
        user_id=user_id,created_at=datetime.now()
        ).order_by('created_at').last()
        print(latest_report)
        return {
        "dashboard_data": report[0],
        "manual_data": {
                "tomorrow_conversation": latest_report.total_expected_conversion if latest_report else "",
                "lead_for_tomorrow": latest_report.actual_expected_conversion if latest_report else "",
                "own_message": latest_report.notes_for_manager if latest_report else "",
        }
        }

        # return report[0]

    except Exception as e:
        raise APIException(e)






def submit_daily_report(user, **data):
    try:
        json_data = data.get("data", {})


        report = DailyReport.objects.create(
            user=user,
            report_date=datetime.now().date(),

            total_leads=json_data.get("total_leads", 0),
            new_leads=json_data.get("new_leads", 0),
            call_spoked=json_data.get("call_spoked", 0),
            not_respond=json_data.get("not_respond", 0),
            follow_up=json_data.get("follow_up", 0),
            pending_follow_up=json_data.get("pending_follow_up", 0),
            partial_payment=json_data.get("partial_payment", 0),
            full_payment=json_data.get("full_payment", 0),
            total_expected_conversion=data.get("tomorrow_conversation", 0),
            actual_expected_conversion=data.get("lead_for_tomorrow", 0),
            notes_for_manager=data.get("own_message", ""),



        #     is_submitted=True,
        #     submitted_at=datetime.now()
        )

        return {
            "status": "success",
            "message": "Report created successfully",
            "data": {
                "id": report.id,
                "report_date": report.report_date,
                "total_leads": report.total_leads,
                "new_leads": report.new_leads,
                "call_spoked": report.call_spoked,
                "not_respond": report.not_respond,
                "follow_up": report.follow_up,
                "pending_follow_up": report.pending_follow_up,
                "partial_payment": report.partial_payment,
                "full_payment": report.full_payment,
                "tomorrow_conversation": report.total_expected_conversion,
                "lead_for_tomorrow": report.actual_expected_conversion,
                "own_message": report.notes_for_manager,
                # "is_submitted": report.is_submitted,
                # "submitted_at": report.submitted_at,
            },
        }

    except Exception as e:
        raise APIException(str(e))




def download_daily_reports(user, **data):
    """API 2: Download report - Get latest report data for today and return response"""
    try:
        today = timezone.now().date()

        # Get the latest report for today's date for this user
        report = (
            DailyReport.objects
            .filter(user_id=user.id, report_date=today)
            .order_by('created_at')   # or 'id' / 'updated_at' - whichever marks "latest"
            .last()
        )

        if not report:
            raise APIException("Report not found for today")
    
        user_details=get_user_details(user)

        # Prepare response with all data
        response_data = {
            'status': 'success',
            'message': 'Report data retrieved successfully',
            'user_name':user_details["name"],
            'user_role': user_details["role"],
            'data': {
                'id': report.id,
                'report_date': report.report_date,
                'user': user.username,
                'total_leads': report.total_leads,
                'new_leads': report.new_leads,
                'call_spoked': report.call_spoked,
                'not_respond': report.not_respond,
                'follow_up': report.follow_up,
                'pending_follow_up': report.pending_follow_up,
                'partial_payment': report.partial_payment,
                'full_payment': report.full_payment,
                "tomorrow_conversation": report.total_expected_conversion,
                "lead_for_tomorrow": report.actual_expected_conversion,
                "own_message": report.notes_for_manager,
                # 'is_submitted': report.is_submitted,
                # 'submitted_at': report.submitted_at,
                'created_at': report.created_at,
                'updated_at': report.updated_at
            }
        }

        return response_data

    except APIException:
        raise
    except Exception as e:
        raise APIException(str(e))
        



