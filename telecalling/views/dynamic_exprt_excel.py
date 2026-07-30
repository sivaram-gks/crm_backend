from rest_framework.views import APIView
from django.http import HttpResponse
from django.apps import apps
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
from ..services.query_services import *
from rest_framework.views import APIView
from rest_framework import serializers
from rest_framework.response import Response
from rest_framework import status
from ..services.lead_services import *

from ..services.payment_services import *

# class DynamicExportExcel(APIView):

#     def get(self, request, app_label, model_name):

#         try:
#             # print(app_label)
#             # print(model_name)
#             # Model get pannrom
#             model = apps.get_model(app_label, model_name)
#         except LookupError:
#             return HttpResponse("Model not found", status=404)

#         # Workbook create
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = model_name

#         # Fields (only DB fields)
#         fields = [field.name for field in model._meta.fields]

#         # Header
#         ws.append(fields)

#         # Data fetch
#         queryset = model.objects.all()

#         for obj in queryset:
#             row = []
#             for field in fields:
#                 value = getattr(obj, field)

#                 # datetime format
#                 if hasattr(value, 'strftime'):
#                     value = value.strftime('%Y-%m-%d %H:%M:%S')

#                 row.append(str(value) if value is not None else '')

#             ws.append(row)

#         # Response
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         filename = f"{model_name}.xlsx"
#         response['Content-Disposition'] = f'attachment; filename={filename}'

#         wb.save(response)
#         return response
    
    
    




class DynamicExportExcel(APIView):

    def post(self, request, *args, **kwargs):
        json_data = request.data.get('data')
        file_name = request.data.get('file_name', 'export')

        if not json_data:
            return HttpResponse("No data provided", status=400)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ✅ Data extract and sheets create
        self.process_data(wb, json_data)

        # ✅ Sheet எதுவும் create ஆகலன்னா
        if not wb.sheetnames:
            ws = wb.create_sheet("Sheet1")
            ws.append(["No Data"])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        wb.save(response)
        return response

    def process_data(self, wb, data, sheet_name="Sheet"):
        """
        எந்த structure வந்தாலும் handle பண்ணும்:
        - list of dicts → direct sheet
        - dict with 'data' key → sheet
        - nested dict → recurse
        """

        # Case 1: List of flat dicts → direct sheet create
        if isinstance(data, list) and data and isinstance(data[0], dict):
            # Check: nested objects இருக்கா இல்லையா
            flat_rows = [row for row in data if not any(isinstance(v, (dict, list)) for v in row.values())]

            if flat_rows:
                self.create_sheet(wb, sheet_name, flat_rows)

            # Nested objects இருந்தா recurse
            for row in data:
                for key, value in row.items():
                    if isinstance(value, (dict, list)):
                        self.process_data(wb, value, key)

        # Case 2: Dict with 'data' key → sheet create
        elif isinstance(data, dict) and 'data' in data:
            rows = data.get('data', [])
            count = data.get('count', None)
            if isinstance(rows, list) and rows:
                self.create_sheet(wb, sheet_name, rows, count)

        # Case 3: Nested dict → each key recurse
        elif isinstance(data, dict):
            for key, value in data.items():
                self.process_data(wb, value, key)

    def create_sheet(self, wb, sheet_name, rows, count=None):
        if not rows:
            return

        clean_name = str(sheet_name).replace('_', ' ').title()[:31]
        base_name = clean_name
        counter = 1
        while clean_name in wb.sheetnames:
            clean_name = f"{base_name[:28]}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=clean_name)

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        start_row = 1

        if count is not None:
            ws.append([f"Total : {count}"])
            count_cell = ws.cell(row=1, column=1)
            count_cell.font = Font(bold=True, color="FFFFFF")
            count_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            # ✅ Count row height
            ws.row_dimensions[1].height = 25
            start_row = 2

        headers = list(rows[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=str(header).replace('_', ' ').upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ✅ Header row height
        ws.row_dimensions[start_row].height = 30

        for row_data in rows:
            ws.append([
                str(row_data.get(col, "")) if row_data.get(col) is not None else ''
                for col in headers
            ])

        # ✅ Data rows height
        total_rows = ws.max_row
        for row_num in range(start_row + 1, total_rows + 1):
            ws.row_dimensions[row_num].height = 20

        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_data in rows:
                cell_val = str(row_data.get(header, "") or "")
                max_length = max(max_length, len(cell_val))
            col_letter = ws.cell(row=start_row, column=col_num).column_letter
            ws.column_dimensions[col_letter].width = max_length + 4
            
            








