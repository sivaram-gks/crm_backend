import csv
import os
import openpyxl
from django.conf import settings
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q
from rest_framework.exceptions import APIException
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from telecalling.models import (
    Lead, PaymentInfo, PaymentHistory, LossLeadDetail, FollowUp,
    CallDetails, CampaignName, LeadSource, PipelineStage,
    User, CoursePlan, CourseName, SelectTag
)

def get_user_display_name(user_obj):
    
    # Safely gets user display name without AnonymousUser errors
    if not user_obj or not getattr(user_obj, 'is_authenticated', False):
        return None
    fname = (getattr(user_obj, 'first_name', '') or "").strip()
    lname = (getattr(user_obj, 'last_name', '') or "").strip()
    
    if fname and lname:
        if fname.lower() == lname.lower():
            return fname
        return f"{fname} {lname}"
    if fname:
        return fname
    return getattr(user_obj, 'username', 'User')


def fetch_all_leads_admin(**data):
    
    # Admin scope version of fetch_leads()
    try:
        now = timezone.now()
        
        base_qs = Lead.objects.select_related(
            "assigned_to",
            "pipeline_stage",
            "campaign",
            "lead_source",
            "course_plan",
            "course_name",
            "course",
        )

        # 1. Search Filter
        search = data.get("search")
        if search:
            base_qs = base_qs.filter(
                Q(full_name__icontains=search) |
                Q(mobile_no__icontains=search) |
                Q(email__icontains=search)
            )

        # 2. Specific Telecaller Filter
        if data.get("tele_id"):
            base_qs = base_qs.filter(assigned_to_id=data.get("tele_id"))

        # 3. Dropdown Filters
        if data.get("pipeline_stage_id"):
            base_qs = base_qs.filter(pipeline_stage_id=data.get("pipeline_stage_id"))
        if data.get("lead_source_id"):
            base_qs = base_qs.filter(lead_source_id=data.get("lead_source_id"))
        if data.get("campaign_name_id"):
            base_qs = base_qs.filter(campaign_id=data.get("campaign_name_id"))
        if data.get("course_plan_id"):
            base_qs = base_qs.filter(course_plan_id=data.get("course_plan_id"))
        if data.get("course_name_id"):
            base_qs = base_qs.filter(course_name_id=data.get("course_name_id"))
        if data.get("priority_id"):
            base_qs = base_qs.filter(priority_id=data.get("priority_id"))

        # 4. Date Filter
        date_filter_type = data.get("date_filter_type") or "all"
        from_date = data.get("from_date")
        to_date = data.get("to_date")
        today = now.date()

        if date_filter_type == "today":
            from_date = to_date = today
        elif date_filter_type == "yesterday":
            from_date = to_date = today - timedelta(days=1)
        elif date_filter_type == "weekly":
            from_date, to_date = today - timedelta(days=7), today
        elif date_filter_type == "monthly":
            from_date, to_date = today.replace(day=1), today

        if from_date and to_date:
            base_qs = base_qs.filter(enquiry_date__date__range=[from_date, to_date])

        # 5. Logic for Tab Classifications (Directly aligned with DB pipeline_stage_id: 1=New, 2=Followup, 3=Won, 4=Lost)
        new_lead_qs = base_qs.filter(Q(pipeline_stage_id=1) | Q(pipeline_stage__name__icontains="new")).distinct()
        follow_up_qs = base_qs.filter(Q(pipeline_stage_id=2) | Q(pipeline_stage__name__icontains="follow")).distinct()
        won_qs = base_qs.filter(Q(pipeline_stage_id=3) | Q(pipeline_stage__name__icontains="won")).distinct()
        lost_qs = base_qs.filter(Q(pipeline_stage_id=4) | Q(pipeline_stage__name__icontains="loss") | Q(pipeline_stage__name__icontains="lost")).distinct()

        # Missed followups: unattended past followups not in main stages
        missed_follow_up_qs = base_qs.filter(
            id__in=FollowUp.objects.filter(is_attended=False, scheduled_at__lt=now).values_list("lead_id", flat=True)
        ).exclude(id__in=new_lead_qs.values_list("id", flat=True)).exclude(id__in=follow_up_qs.values_list("id", flat=True)).exclude(id__in=won_qs.values_list("id", flat=True)).exclude(id__in=lost_qs.values_list("id", flat=True)).distinct()

        tabs = {
            "all": base_qs,
            "new_lead": new_lead_qs,
            "new": new_lead_qs,
            "follow_up": follow_up_qs,
            "missed_follow_up": missed_follow_up_qs,
            "pending_follow_up": missed_follow_up_qs,
            "won": won_qs,
            "lost": lost_qs,
            "loss": lost_qs,
        }

        # Stats Counts matching UI Screen 100%:
        stats = {
            "total_count": tabs["all"].count(),
            "new_count": tabs["new_lead"].count(),           # ➔ 33
            "follow_up_count": tabs["follow_up"].count(),     # ➔ 19
            "pending_follow_up_count": tabs["missed_follow_up"].count(), # ➔ 0
            "won_count": tabs["won"].count(),               # ➔ 17
            "loss_count": tabs["lost"].count(),             # ➔ 6
        }

        tab_counts = {key: qs.count() for key, qs in tabs.items()}

        # 6. Selected Tab & Pagination
        lead_filter_type = data.get("lead_filter_type") or "all"
        selected_qs = tabs.get(lead_filter_type, tabs["all"]).order_by(data.get("sort_by") or "-created_at")

        total = selected_qs.count()
        page_size_input = data.get("page_size")
        # page_size = 0, "all", அல்லது 1000 என அனுப்பினால் அத்தனை 63+ லீட்களையும் ஒரே பக்கத்தில் தரும்
        
        if str(page_size_input).lower() in ["0", "all", "none"] or page_size_input == 0:
            page = 1
            page_size = total
            start = 0 
            rows = selected_qs
        else:
            page = int(data.get("page") or 1)
            page_size = int(page_size_input or 1000)  # 👈 Default limit set to 1000
            start = (page - 1) * page_size
            end = start + page_size
            rows = selected_qs[start:end]


        leads = []
        for idx, lead in enumerate(rows, start=start + 1):
            latest_call = CallDetails.objects.filter(lead=lead).order_by('-created_at').first()
            latest_followup = FollowUp.objects.filter(lead=lead, is_attended=False).order_by('scheduled_at').first()
            payment_info = PaymentInfo.objects.filter(lead=lead).first()
            
            course_fee = lead.course.course_fees if (lead.course and hasattr(lead.course, 'course_fees')) else 0
            if not course_fee and payment_info:
                course_fee = (payment_info.amount_paid or 0) + (payment_info.pending_amount or 0)
            
            pending_amt = payment_info.pending_amount if payment_info else 0

            leads.append({
                # "s_no": idx,
                "id": lead.id,
                "full_name": lead.full_name or "",
                "mobile_no": lead.mobile_no or "",
                "assigned_to_id": lead.assigned_to_id,
                "assigned_to": get_user_display_name(lead.assigned_to),
                "stage": lead.pipeline_stage.name if lead.pipeline_stage else "New",
                "pipeline": "Education",
                "campaign": lead.campaign.name if lead.campaign else None,
                "source": lead.lead_source.name if lead.lead_source else None,
                "course_plan": lead.course_plan.courseplan if lead.course_plan else None,
                "course": lead.course_name.coursename if lead.course_name else None,
                "next_followup": latest_followup.scheduled_at if latest_followup else None,
                "amount": course_fee,
                "pending_amount": pending_amt,
                "last_contacted": latest_call.called_at if latest_call else None,
                "last_conversation_outcome": latest_call.conversation_summary if latest_call else None,
                "created": lead.created_at or lead.enquiry_date,
                "enquiry_date": lead.enquiry_date,
            })

        return {
            "leads": leads,
            "stats": stats,   
            "tab_counts": tab_counts,
            "page": page,
            "page_size": page_size,
            "total": total
        }

    except Exception as e:
        raise APIException(str(e))


# ------------------------------------- admin add new lead services ------------------------------------------

def get_user_display_name(user_obj):
    """Safely gets user display name without AnonymousUser errors"""
    if not user_obj or not getattr(user_obj, 'is_authenticated', False):
        return None
    fname = (getattr(user_obj, 'first_name', '') or "").strip()
    lname = (getattr(user_obj, 'last_name', '') or "").strip()
    
    if fname and lname:
        if fname.lower() == lname.lower():
            return fname
        return f"{fname} {lname}"
    if fname:
        return fname
    return getattr(user_obj, 'username', 'User')


def get_add_lead_dropdowns_admin():
    """
    Returns all dropdown options for Add New Lead modal:
    1. pipelines: Education, Product
    2. campaigns: All rows from CampaignName table
    3. sources: All rows from LeadSource table
    4. telecallers: All active Telecallers
    """
    try:
        pipelines = [
            {"id": "Education", "name": "Education"},
            {"id": "Product", "name": "Product"}
        ]
        campaigns_qs = CampaignName.objects.all()
        campaigns = [{"id": c.id, "name": c.name} for c in campaigns_qs]
        sources_qs = LeadSource.objects.all()
        sources = [{"id": s.id, "name": s.name} for s in sources_qs]
        users_qs = User.objects.filter(is_active=True)
        telecallers = [{"id": u.id, "name": get_user_display_name(u)} for u in users_qs]
        return {
            "pipelines": pipelines,
            "campaigns": campaigns,
            "sources": sources,
            "telecallers": telecallers
        }
    except Exception as e:
        raise APIException(str(e))
    
    
def add_new_lead_admin(user, **data):
    """
    Creates a new lead with AnonymousUser safety and Figma specs.
    """
    try:
        mobile_no = data.get("mobile_no") or data.get("mobile")
        if not mobile_no:
            raise APIException("Mobile number is required")
        
        # 1. Duplicate Check
        existing_lead = Lead.objects.filter(mobile_no=mobile_no).select_related('assigned_to').first()
        if existing_lead:
            assigned_name = get_user_display_name(existing_lead.assigned_to) or "another agent"
            raise APIException(f"Mobile number {mobile_no} is already registered and assigned to {assigned_name}.")
        
        # 2. Name Combination
        first_name = (data.get("first_name") or "").strip()
        last_name = (data.get("last_name") or "").strip()
        
        if first_name or last_name:
            full_name = f"{first_name} {last_name}".strip()
        else:
            full_name = (data.get("full_name") or "New Enquiry").strip()
            
        # 3. Campaign (by ID or Name)
        campaign_id = data.get("campaign_id") or data.get("campaign_name_id")
        campaign_name_str = data.get("campaign") or data.get("campaign_name")
        
        campaign = None
        if campaign_id:
            campaign = CampaignName.objects.filter(id=campaign_id).first()
        elif campaign_name_str:
            campaign = CampaignName.objects.filter(name__icontains=str(campaign_name_str).strip()).first()
            
        if not campaign:
            campaign = CampaignName.objects.first()
            
        # 4. Lead Source (by ID or Name)
        source_id = data.get("lead_source_id") or data.get("source_id")
        source_name_str = data.get("source") or data.get("lead_source")
        
        source = None
        if source_id:
            source = LeadSource.objects.filter(id=source_id).first()
        elif source_name_str:
            source = LeadSource.objects.filter(name__icontains=str(source_name_str).strip()).first()
            
        if not source:
            source = LeadSource.objects.first()
            
        # 5. Pipeline Name
        pipeline_name = str(data.get("pipeline") or "Education").strip()
        stage_id = data.get("pipeline_stage_id") or 1
        stage = PipelineStage.objects.filter(id=stage_id).first()
        if not stage:
            stage = PipelineStage.objects.filter(name__icontains="new").first()
            
        # 6. Assigned Telecaller User
        assigned_to_id = data.get("assigned_to_id")
        if not assigned_to_id and user and getattr(user, 'is_authenticated', False):
            assigned_to_id = user.id
            
        # 7. Safe Creator Name Check
        if user and getattr(user, 'is_authenticated', False):
            user_role = "Admin" if (getattr(user, 'is_superuser', False) or str(getattr(user, 'user_type', '')).lower() == 'admin') else "Telecaller"
            creator_name = get_user_display_name(user) or getattr(user, 'username', 'User')
            created_by_info = f"{creator_name} ({user_role})"
        else:
            created_by_info = "Admin API (Unauthenticated)"
            
        # 8. Create Lead Row
        lead = Lead.objects.create(
            full_name=full_name,
            mobile_no=mobile_no,
            email=data.get("email"),
            campaign=campaign,
            lead_source=source,
            pipeline_stage=stage,
            assigned_to_id=assigned_to_id,
            enquiry_date=data.get("enquiry_date") or timezone.now(),
            current_status="working",
            priority_id=data.get("priority_id") or 4,
            created_by=created_by_info
        )
        return {
            "status": "success",
            "message": f"Lead '{lead.full_name}' created successfully!",
            "lead_id": lead.id,
            "pipeline": pipeline_name,
            "created_by": created_by_info,
            "assigned_to_id": lead.assigned_to_id
        }
    except Exception as e:
        raise APIException(str(e))
    
    
    
    
# -----------------------------upload lead excel file service------------------------------------------   
    
def upload_lead_excel_admin(file_obj, user=None):
    """
    Bulk Excel/CSV Upload Service matching Figma Modal specs:
    - Supported formats: .csv, .xls, .xlsx
    - File size limit: 3MB
    - Max rows: 25,000
    - Auto maps columns (name, mobile, email, campaign, source)
    - Skips duplicate mobile numbers
    """
    try:
        if not file_obj:
            raise APIException("Please select an Excel or CSV file to upload.")
        
        # 1. File Size Validation (Max 3MB = 3 * 1024 * 1024 bytes)
        if file_obj.size > 3 * 1024 * 1024:
            raise APIException("File size exceeds the 3MB limit. Please upload a file smaller than 3MB.")
        
        # 2. File Extension Validation
        filename = file_obj.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            raise APIException("Unsupported file format! Please upload a .csv, .xls, or .xlsx file.")
        rows_data = []

        # 3. Reading File Data (.csv vs .xlsx)
        if filename.endswith('.csv'):
            decoded_file = file_obj.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file)
            for row in reader:
                if any(row):
                    rows_data.append(row)
        else:  # .xlsx or .xls
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if row and any(row):
                    rows_data.append([str(cell) if cell is not None else "" for cell in row])
        if not rows_data or len(rows_data) < 2:
            raise APIException("The uploaded file is empty or missing data rows.")
        
        # Row limit check (Max 25,000 rows as per Figma)
        data_rows = rows_data[1:]
        if len(data_rows) > 25000:
            raise APIException("Maximum limit of 25,000 leads exceeded per file.")
        
        # 4. Header Column Mapping
        headers = [str(h).strip().lower() for h in rows_data[0]]
        
        name_idx = -1
        mobile_idx = -1
        email_idx = -1
        campaign_idx = -1
        source_idx = -1
        for idx, h in enumerate(headers):
            if any(k in h for k in ['name', 'full_name', 'first_name', 'student']):
                if name_idx == -1: name_idx = idx
            elif any(k in h for k in ['mobile', 'phone', 'contact', 'number']):
                if mobile_idx == -1: mobile_idx = idx
            elif 'email' in h:
                if email_idx == -1: email_idx = idx
            elif 'campaign' in h:
                if campaign_idx == -1: campaign_idx = idx
            elif 'source' in h:
                if source_idx == -1: source_idx = idx
       
        # Fallback to column indices if headers not named explicitly
        if mobile_idx == -1:
            mobile_idx = 1 if len(headers) > 1 else 0
        if name_idx == -1:
            name_idx = 0
       
        # Default Foreign Keys
        default_stage = PipelineStage.objects.filter(id=1).first() or PipelineStage.objects.first()
        default_campaign = CampaignName.objects.first()
        default_source = LeadSource.objects.first()
        default_agent = User.objects.filter(is_active=True).first()
        
        # Creator Info
        creator_info = "Admin Excel Upload"
        if user and getattr(user, 'is_authenticated', False):
            creator_info = f"{getattr(user, 'username', 'User')} (Admin Excel)"
        success_count = 0
        duplicate_count = 0
        duplicate_mobiles = []
        
        # 5. Loop Through Rows & Insert Leads into telecalling_lead Table
        for row in data_rows:
            full_name = str(row[name_idx]).strip() if name_idx < len(row) else "New Enquiry"
            mobile_no = str(row[mobile_idx]).strip() if mobile_idx < len(row) else ""
            
            # Clean mobile number format
            mobile_no = mobile_no.replace(".0", "").strip()
            if not mobile_no:
                continue
            email = str(row[email_idx]).strip() if (email_idx != -1 and email_idx < len(row)) else None
           
            # Duplicate Mobile Check
            if Lead.objects.filter(mobile_no=mobile_no).exists():
                duplicate_count += 1
                duplicate_mobiles.append(mobile_no)
                continue
            
            # Dynamic Campaign & Source Resolution
            campaign_val = str(row[campaign_idx]).strip() if (campaign_idx != -1 and campaign_idx < len(row)) else ""
            campaign_obj = CampaignName.objects.filter(name__icontains=campaign_val).first() if campaign_val else default_campaign
            source_val = str(row[source_idx]).strip() if (source_idx != -1 and source_idx < len(row)) else ""
            source_obj = LeadSource.objects.filter(name__icontains=source_val).first() if source_val else default_source
            
            # Create Lead Row in Database
            Lead.objects.create(
                full_name=full_name or "New Enquiry",
                mobile_no=mobile_no,
                email=email,
                campaign=campaign_obj or default_campaign,
                lead_source=source_obj or default_source,
                pipeline_stage=default_stage,
                assigned_to=default_agent,
                current_status="working",
                priority_id=4,
                created_by=creator_info,
                enquiry_date=timezone.now()
            )
            success_count += 1
        return {
            "status": "success",
            "message": f"{success_count} leads uploaded successfully! {duplicate_count} duplicate mobile numbers skipped.",
            "total_rows": len(data_rows),
            "success_count": success_count,
            "duplicate_count": duplicate_count,
            "skipped_duplicates": duplicate_mobiles[:10]
        }
    except Exception as e:
        raise APIException(str(e))
    
    
    
    
# --------------------------------export leads to excel service------------------------------------------

def export_all_leads_admin(**data):
    """
    Admin Leads Page -> Export to Excel (.xlsx) Service.
    Lime Green Header Styling (#84C225) & Spacious Column Widths matching reference image.
    """
    try:
        # 1. Normalize filter type
        raw_filter = (
            data.get("lead_filter_type") or 
            data.get("filter_type") or 
            data.get("stage") or 
            data.get("tab") or 
            "all"
        )
        lead_filter_type = str(raw_filter).lower().strip()
        
        data["lead_filter_type"] = lead_filter_type
        data['page_size'] = "all"

        # 2. Fetch matching leads from fetch_all_leads_admin
        result = fetch_all_leads_admin(**data)
        leads = result.get("leads", [])

        # 3. Create Excel Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Leads_{lead_filter_type}"

        # 🎨 4. Exact 14 Column Headers matching reference image
        headers = [
            "s_no", 
            "id", 
            "full_name", 
            "mobile_no", 
            "email", 
            "tag", 
            "stage", 
            "source", 
            "campaign_name", 
            "course_plan", 
            "course_name", 
            "pending_amount", 
            "total_amount", 
            "created_at"
        ]
        ws.append(headers)

        # 5. Populate Data Rows
        for idx, lead in enumerate(leads, start=1):
            ws.append([
                idx,                                                                   # s_no
                lead.get("id"),                                                        # id
                lead.get("full_name") or "",                                           # full_name
                lead.get("mobile_no") or "",                                           # mobile_no
                lead.get("email") or "",                                              # email
                lead.get("tag") or "new",                                              # tag
                lead.get("stage") or "",                                               # stage
                lead.get("source") or "",                                              # source
                lead.get("campaign") or "",                                            # campaign_name
                lead.get("course_plan") or "",                                         # course_plan
                lead.get("course") or "",                                              # course_name
                lead.get("pending_amount") or 0,                                       # pending_amount
                lead.get("amount") or 0,                                               # total_amount
                str(lead.get("created")) if lead.get("created") else ""                # created_at
            ])

        # 🎨 6. LIME GREEN HEADER STYLING (#84C225) & BORDERS
        header_fill = PatternFill(start_color="84C225", end_color="84C225", fill_type="solid")  # Bright Lime Green
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")                   # Bold White Text
        data_font = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

       # Apply Header Styling (Center Aligned)
        ws.row_dimensions[1].height = 26
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # 🎯 Apply Data Rows Styling (Neat Center Alignment for ALL Cells)
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 22
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align

        # 📐 7. AUTO COLUMN WIDTH (Spacious Display)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 6, 16)

        # 8. Save & Return File Link
        file_name = f"Admin_Leads_{lead_filter_type}.xlsx"
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, file_name)

        wb.save(file_path)

        return {
            "status": "success",
            "message": f"Successfully exported {len(leads)} leads for tab '{lead_filter_type}'!",
            "total_exported": len(leads),
            "lead_filter_type": lead_filter_type,
            "file_name": file_name,
            "download_url": f"/media/exports/{file_name}"
        }

    except Exception as e:
        raise APIException(str(e))
    
    
    
# ----------------------------get_filter_dropdowns_admin----------------------------

def get_filter_dropdowns_admin():
    """
    Admin Leads Page -> Filter Modal Dropdown Options API.
    Returns dynamic options for:
    1. Pipeline Stage
    2. Lead Source
    3. Campaign Name
    4. Course Plan
    5. Assigned User (Telecallers)
    """
    try:
        pipeline_stages_qs = PipelineStage.objects.all()
        pipeline_stages = [{"id": p.id, "name": p.name} for p in pipeline_stages_qs]

        lead_sources_qs = LeadSource.objects.all()
        lead_sources = [{"id": s.id, "name": s.name} for s in lead_sources_qs]

        campaigns_qs = CampaignName.objects.all()
        campaigns = [{"id": c.id, "name": c.name} for c in campaigns_qs]

        course_plans_qs = CoursePlan.objects.all()
        course_plans = [{"id": cp.id, "name": getattr(cp, 'courseplan', getattr(cp, 'name', str(cp)))} for cp in course_plans_qs]

        users_qs = User.objects.filter(is_active=True)
        telecallers = [{"id": u.id, "name": get_user_display_name(u)} for u in users_qs]

        return {
            "pipeline_stages": pipeline_stages,
            "lead_sources": lead_sources,
            "campaigns": campaigns,
            "course_plans": course_plans,
            "telecallers": telecallers
        }
    except Exception as e:
        raise APIException(str(e))
    
    
    
    
# ------------------------------------- fetch_pipeline_leads_admin----------------------------------

def fetch_pipeline_leads_admin(**data):
    """
    Admin Pipeline View (Kanban Cards API).
    Returns exact JSON structure matching Figma Kanban UI with 5 Columns:
    1. new_lead: Array of cards
    2. follow_up: Dict with total_count, past, current, future
    3. unreached_calls: Dict with total_count, past, current, future
    4. pending_payment: Dict with total_count, past, current, future
    5. closed: Dict with total_count, no_response, not_reachable, wrong_number, won, lost
    """
    try:
        now = timezone.now()
        today = now.date()

        base_qs = Lead.objects.select_related(
            "assigned_to",
            "pipeline_stage",
            "campaign",
            "lead_source"
        )

        # 1. Search Filter
        search = data.get("search")
        if search:
            base_qs = base_qs.filter(
                Q(full_name__icontains=search) |
                Q(mobile_no__icontains=search) |
                Q(email__icontains=search)
            )

        # 2. Assigned Telecaller Filter
        tele_id = data.get("assigned_to_id") or data.get("tele_id")
        if tele_id and int(tele_id) > 0:
            base_qs = base_qs.filter(assigned_to_id=tele_id)

        # 3. Dropdown Filters
        if data.get("pipeline_stage_id") and int(data.get("pipeline_stage_id")) > 0:
            base_qs = base_qs.filter(pipeline_stage_id=data.get("pipeline_stage_id"))
        if data.get("lead_source_id") and int(data.get("lead_source_id")) > 0:
            base_qs = base_qs.filter(lead_source_id=data.get("lead_source_id"))
        if data.get("campaign_name_id") and int(data.get("campaign_name_id")) > 0:
            base_qs = base_qs.filter(campaign_id=data.get("campaign_name_id"))
        if data.get("course_plan_id") and int(data.get("course_plan_id")) > 0:
            base_qs = base_qs.filter(course_plan_id=data.get("course_plan_id"))

        # 4. Date Filter
        date_filter_type = data.get("date_filter_type") or "all"
        from_date = data.get("from_date")
        to_date = data.get("to_date")

        if date_filter_type == "today":
            from_date = to_date = today
        elif date_filter_type == "yesterday":
            from_date = to_date = today - timedelta(days=1)
        elif date_filter_type == "weekly":
            from_date, to_date = today - timedelta(days=7), today
        elif date_filter_type == "monthly":
            from_date, to_date = today.replace(day=1), today

        if from_date and to_date and str(from_date).strip() != "" and str(to_date).strip() != "":
            base_qs = base_qs.filter(enquiry_date__date__range=[from_date, to_date])

        # Card Formatting Helper (Exact UI Match: e.g. "31 Jan, 10:55 AM")
        def format_card(lead):
            created_dt = lead.enquiry_date or lead.created_at
            formatted_date = created_dt.strftime("%d %b, %I:%M %p") if created_dt else ""
            return {
                "id": lead.id,
                "full_name": lead.full_name or "",
                "mobile_no": lead.mobile_no or "",
                "assigned_to": get_user_display_name(lead.assigned_to) or "Unassigned",
                "source": lead.lead_source.name if lead.lead_source else "direct walk in",
                "created_at": formatted_date
            }

        # 1. New Lead Column (Stage ID 1)
        new_lead_qs = base_qs.filter(Q(pipeline_stage_id=1) | Q(pipeline_stage__name__icontains="new")).order_by("-id")
        new_lead_cards = [format_card(l) for l in new_lead_qs]

        # 2. Follow Up Column (Stage ID 2)
        follow_up_qs = base_qs.filter(Q(pipeline_stage_id=2) | Q(pipeline_stage__name__icontains="follow")).order_by("-id")
        follow_up_cards = [format_card(l) for l in follow_up_qs]

        follow_up_dict = {
            "total_count": len(follow_up_cards),
            "past": [],
            "current": follow_up_cards,
            "future": []
        }

        # 3. Unreached Calls Column
        unreached_dict = {
            "total_count": 0,
            "past": [],
            "current": [],
            "future": []
        }

        # 4. Pending Payment Column
        # Database-ல் நிலுவைத் தொகை உள்ள லீட்களைப் (Pending Payment) படித்தல்:
        pending_pay_lead_ids = PaymentInfo.objects.filter(pending_amount__gt=0).values_list("lead_id", flat=True)
        pending_payment_qs = base_qs.filter(id__in=pending_pay_lead_ids).order_by("-id")
        pending_payment_cards = [format_card(l) for l in pending_payment_qs]
        pending_payment_dict = {
            "total_count": len(pending_payment_cards),
            "past": [],
            "current": pending_payment_cards,
            "future": []
        }

        # 5. Closed Column (Won=Stage ID 3, Lost=Stage ID 4)
        won_qs = base_qs.filter(Q(pipeline_stage_id=3) | Q(pipeline_stage__name__icontains="won")).order_by("-id")
        lost_qs = base_qs.filter(Q(pipeline_stage_id=4) | Q(pipeline_stage__name__icontains="loss") | Q(pipeline_stage__name__icontains="lost")).order_by("-id")

        won_cards = [format_card(l) for l in won_qs]
        lost_cards = [format_card(l) for l in lost_qs]

        closed_dict = {
            "total_count": len(won_cards) + len(lost_cards),
            "no_response": [],
            "not_reachable": [],
            "wrong_number": [],
            "won": won_cards,
            "lost": lost_cards
        }

        return {
            "status": True,
            "message": "Pipeline leads fetched successfully",
            "data": {
                "new_lead": new_lead_cards,
                "follow_up": follow_up_dict,
                "unreached_calls": unreached_dict,
                "pending_payment": pending_payment_dict,
                "closed": closed_dict
            }
        }

    except Exception as e:
        raise APIException(str(e))

# ---------------------------- fetch lead and timeline details for admin ------------------------------------------

def fetch_lead_details_admin(**data):
    """
    Admin Lead Details Modal & Activity Timeline History API.
    Returns:
    1. lead_info: Summary details for left card
    2. timeline: Complete chronological activity history for right timeline
    """
    try:
        lead_id = data.get("lead_id")
        if not lead_id:
            raise APIException("Lead ID is required")

        lead = Lead.objects.select_related(
            "assigned_to",
            "pipeline_stage",
            "campaign",
            "lead_source",
            "course_plan",
            "course_name",
            "course",
            "priority"
        ).filter(id=lead_id).first()

        if not lead:
            raise APIException("Lead not found")

        # 1. Latest Call & Followup
        latest_call = CallDetails.objects.filter(lead=lead).order_by('-created_at').first()
        latest_followup = FollowUp.objects.filter(lead=lead, is_attended=False).order_by('scheduled_at').first()
        payment_info = PaymentInfo.objects.filter(lead=lead).first()

        course_fee = lead.course.course_fees if (lead.course and hasattr(lead.course, 'course_fees')) else 0
        if not course_fee and payment_info:
            course_fee = (payment_info.amount_paid or 0) + (payment_info.pending_amount or 0)

        # 2. Left Card Lead Info Summary
        lead_info = {
            "id": lead.id,
            "full_name": lead.full_name or "",
            "mobile_no": lead.mobile_no or "",
            "email": lead.email or "",
            "alternative_mobile": lead.alternative_mobile or "",
            "location": lead.location or "",
            "passed_out_year": lead.passed_out_year,
            "experience": lead.experience or "",
            "stage": lead.pipeline_stage.name if lead.pipeline_stage else "New Lead",
            "pipeline": "Education",
            "campaign": lead.campaign.name if lead.campaign else None,
            "source": lead.lead_source.name if lead.lead_source else None,
            "course_plan": lead.course_plan.courseplan if lead.course_plan else None,
            "course": lead.course_name.coursename if lead.course_name else None,
            "assigned_to": get_user_display_name(lead.assigned_to) or "Unassigned",
            "created": lead.enquiry_date.strftime("%Y-%m-%d") if lead.enquiry_date else str(lead.created_at.date()) if lead.created_at else "",
            "last_contacted": latest_call.created_at.strftime("%d-%m-%Y") if (latest_call and latest_call.created_at) else None,
            "next_followup": latest_followup.scheduled_at.strftime("%d-%m-%Y") if (latest_followup and latest_followup.scheduled_at) else None,
            "latest_conversation_outcome": latest_call.conversation_summary if latest_call else None,
            "amount": course_fee,
            "pending_amount": payment_info.pending_amount if payment_info else 0
        }

        # 3. Right Card Timeline Events
        timeline = []

        # A. Call Events from CallDetails
        calls = CallDetails.objects.filter(lead=lead).select_related("stage", "select_tag", "telecaller").order_by("-created_at")
        for c in calls:
            conn_status = str(c.connection_status or "Connected").lower()
            tag_name = c.select_tag.name if c.select_tag else None
            stage_name = c.stage.name if c.stage else None

            # Type mapping for UI badges
            if "not" in conn_status or "unreach" in conn_status:
                event_type = "not_connected"
                event_title = "Not Connected"
            else:
                event_type = "lead_disposed" if (stage_name or tag_name) else "answered"
                event_title = f"Lead Disposed | {conn_status.capitalize()}" if (stage_name or tag_name) else "Answered"

            formatted_time = c.created_at.strftime("%I:%M %p") if c.created_at else ""
            formatted_date = c.created_at.strftime("%d %b %Y") if c.created_at else ""

            timeline.append({
                "id": f"call_{c.id}",
                "date": formatted_date,
                "time": formatted_time,
                "type": event_type,
                "title": event_title,
                "stage_tag": stage_name,
                "sub_tag": tag_name,
                "remark": c.conversation_summary or "",
                "recording_url": c.upload_recording.url if (c.upload_recording and hasattr(c.upload_recording, 'url')) else None,
                "duration_seconds": c.duration_seconds or 0,
                "created_by": get_user_display_name(c.telecaller) or c.created_by
            })

        # B. Loss Details (if lead was lost)
        loss_detail = LossLeadDetail.objects.filter(lead=lead).select_related("main_reason").first()
        if loss_detail:
            reason_name = loss_detail.main_reason.name if loss_detail.main_reason else "Closed"
            formatted_date = loss_detail.created_at.strftime("%d %b %Y") if loss_detail.created_at else ""
            formatted_time = loss_detail.created_at.strftime("%I:%M %p") if loss_detail.created_at else ""
            timeline.append({
                "id": f"loss_{loss_detail.id}",
                "date": formatted_date,
                "time": formatted_time,
                "type": "closed",
                "title": "Lead Closed",
                "stage_tag": "Closed",
                "sub_tag": reason_name,
                "remark": loss_detail.detailed_reason or "",
                "recording_url": None,
                "created_by": loss_detail.created_by
            })

        # C. Lead Creation Event
        created_dt = lead.enquiry_date or lead.created_at
        if created_dt:
            timeline.append({
                "id": "created",
                "date": created_dt.strftime("%d %b %Y"),
                "time": created_dt.strftime("%I:%M %p"),
                "type": "created",
                "title": "Lead Created",
                "stage_tag": None,
                "sub_tag": None,
                "remark": f"Source: {lead.lead_source.name if lead.lead_source else 'Direct'}",
                "recording_url": None,
                "created_by": lead.created_by or "Admin"
            })

        return {
            "status": "success",
            "data": {
                "lead_info": lead_info,
                "timeline": timeline
            }
        }

    except Exception as e:
        raise APIException(str(e))
    
    
    
# ----------------------------get_mark_as_won_info_admin--------------------------

def get_mark_as_won_info_admin(lead_id):
    """
    Mark as Won Modal Open -> Fetch lead summary & payment info.
    Safe handling when lead.course is None.
    """
    try:
        lead = Lead.objects.select_related("assigned_to", "pipeline_stage", "course", "course_name").filter(id=lead_id).first()
        if not lead:
            raise APIException("Lead not found")

        payment_info = PaymentInfo.objects.filter(lead=lead).first()
        
        # 🛡️ Safe Course Fee Check (Prevents NoneType error if lead.course is None)
        course_fee = 16000
        if lead.course and getattr(lead.course, 'course_fees', None):
            course_fee = lead.course.course_fees

        paid_amount = payment_info.amount_paid if payment_info else 0
        pending_amount = max(course_fee - paid_amount, 0)

        lead_info = {
            "lead_id": lead.id,
            "full_name": lead.full_name or "",
            "mobile_no": lead.mobile_no or "",
            "assigned_to": get_user_display_name(lead.assigned_to) or "Unassigned",
            "current_stage": lead.pipeline_stage.name if lead.pipeline_stage else "Hot",
            "course_fee": course_fee,
            "amount_paid": paid_amount,
            "pending_amount": pending_amount
        }

        payment_modes = ["Online", "Cash", "UPI", "Bank Transfer", "Cheque"]
        lead_stages = ["Hot", "Warm", "Cold"]

        return {
            "status": "success",
            "data": {
                "lead_info": lead_info,
                "payment_modes": payment_modes,
                "lead_stages": lead_stages
            }
        }
    except Exception as e:
        raise APIException(str(e))


def mark_as_won_admin(**data):
    """
    Submit Mark as Won Modal -> Update Lead to WON & Record Payment Details.
    100% Safe course fee calculation preventing NoneType AttributeError.
    """
    try:
        lead_id = data.get("lead_id")
        if not lead_id:
            raise APIException("Lead ID is required")

        lead = Lead.objects.select_related("course").filter(id=lead_id).first()
        if not lead:
            raise APIException("Lead not found")

        paid_through = data.get("paid_through") or "Online"
        amount_paid = float(data.get("amount_paid") or 0)
        is_full_payment = bool(data.get("is_full_payment", False))
        due_date = data.get("due_date")
        summary = data.get("summary") or ""
        next_followup = data.get("next_followup")

        # 1. Update Lead Pipeline Stage to WON (Stage ID 3)
        won_stage = PipelineStage.objects.filter(id=3).first() or PipelineStage.objects.filter(name__icontains="won").first()
        if won_stage:
            lead.pipeline_stage = won_stage
            lead.save()

        # 🛡️ 2. Safe Course Fee Calculation (Prevents NoneType error if lead.course is None)
        course_fee = 16000
        if lead.course and getattr(lead.course, 'course_fees', None):
            course_fee = lead.course.course_fees

        if is_full_payment:
            amount_paid = course_fee
            pending_amount = 0
        else:
            pending_amount = max(course_fee - amount_paid, 0)

        # 3. Create or Update PaymentInfo
        payment_obj, created = PaymentInfo.objects.get_or_create(
            lead=lead,
            defaults={
                'amount_paid': amount_paid,
                'pending_amount': pending_amount,
                'is_full_payment': (pending_amount == 0),
                'summary': summary
            }
        )
        if not created:
            payment_obj.amount_paid = (payment_obj.amount_paid or 0) + amount_paid
            payment_obj.pending_amount = pending_amount
            payment_obj.is_full_payment = (pending_amount == 0)
            payment_obj.summary = summary
            payment_obj.save()

        # 4. Create PaymentHistory Record
        PaymentHistory.objects.create(
            payment=payment_obj,
            paid_amount=amount_paid,
            pending_amount=pending_amount,
            notes=f"Paid via {paid_through}. {summary}",
            due_date=due_date if due_date else None
        )

        # 5. Create FollowUp if next_followup date is provided
        if next_followup and str(next_followup).strip() != "":
            FollowUp.objects.create(
                lead=lead,
                telecaller=lead.assigned_to or User.objects.filter(is_active=True).first(),
                scheduled_at=next_followup,
                notes=f"Followup after Won ({summary})",
                is_attended=False
            )

        return {
            "status": "success",
            "message": f"Lead '{lead.full_name}' marked as WON successfully!",
            "lead_id": lead.id,
            "stage": "Won",
            "amount_paid": payment_obj.amount_paid,
            "pending_amount": payment_obj.pending_amount
        }

    except Exception as e:
        raise APIException(str(e))
    
    
    
# -----------------------------get_mark_as_lost_info_admin----------------------

def get_mark_as_lost_info_admin(lead_id):
    """
    Mark as Lost Modal Open -> Fetch lead summary (attempts, last contacted, lead age) & loss reasons.
    """
    try:
        lead = Lead.objects.select_related("assigned_to", "pipeline_stage").filter(id=lead_id).first()
        if not lead:
            raise APIException("Lead not found")

        calls = CallDetails.objects.filter(lead=lead).order_by("-created_at")
        total_attempts = calls.count()
        latest_call = calls.first()

        # Lead Age Calculation (e.g., "3 Months" or "15 Days")
        created_dt = lead.enquiry_date or lead.created_at
        if created_dt:
            delta_days = (timezone.now() - created_dt).days
            if delta_days >= 30:
                lead_age = f"{delta_days // 30} Months"
            else:
                lead_age = f"{delta_days} Days"
        else:
            lead_age = "1 Month"

        last_contacted_str = latest_call.created_at.strftime("%d %b, %I:%M %p") if (latest_call and latest_call.created_at) else "No calls yet"
        last_conversation_str = latest_call.conversation_summary if latest_call else "No conversation recorded"
        attempts_str = f"{total_attempts} Calls done ({last_contacted_str})"

        lead_info = {
            "lead_id": lead.id,
            "full_name": lead.full_name or "",
            "mobile_no": lead.mobile_no or "",
            "assigned_to": get_user_display_name(lead.assigned_to) or "Unassigned",
            "total_attempts": attempts_str,
            "last_conversation": last_conversation_str,
            "last_contacted": last_contacted_str,
            "lead_age": lead_age,
            "current_stage": lead.pipeline_stage.name if lead.pipeline_stage else "Cold"
        }

        # Loss Reasons from SelectTag model
        tags = SelectTag.objects.all()
        main_reasons = [{"id": t.id, "name": t.name} for t in tags]

        return {
            "status": "success",
            "data": {
                "lead_info": lead_info,
                "main_reasons": main_reasons,
                "sub_reasons": main_reasons
            }
        }
    except Exception as e:
        raise APIException(str(e))


def mark_as_lost_admin(**data):
    """
    Submit Mark as Lost Modal -> Update Lead to LOST (Stage ID 4) & Record Loss Details.
    Robustly handles both string names and integer IDs from frontend.
    """
    try:
        lead_id = data.get("lead_id") or data.get("id")
        if not lead_id:
            raise APIException("Lead ID is required")

        lead = Lead.objects.filter(id=lead_id).first()
        if not lead:
            raise APIException("Lead not found")

        main_reason_val = data.get("main_reason_id") or data.get("main_reason") or data.get("reason")
        detailed_reason = data.get("detailed_reason") or data.get("remarks") or data.get("reason") or ""

        # 1. Update Lead Pipeline Stage to LOST (Stage ID 4)
        lost_stage = PipelineStage.objects.filter(id=4).first() or PipelineStage.objects.filter(Q(name__icontains="loss") | Q(name__icontains="lost")).first()
        if lost_stage:
            lead.pipeline_stage = lost_stage
            lead.save()

        # 2. Resolve Main Reason Tag (Handles both integer IDs and string reason names)
        main_reason_obj = None
        if main_reason_val:
            if str(main_reason_val).isdigit():
                main_reason_obj = SelectTag.objects.filter(id=int(main_reason_val)).first()
            else:
                main_reason_obj = SelectTag.objects.filter(Q(name__icontains=str(main_reason_val))).first()

        # 3. Create or Update LossLeadDetail
        loss_obj, created = LossLeadDetail.objects.get_or_create(
            lead=lead,
            defaults={
                "main_reason": main_reason_obj,
                "detailed_reason": detailed_reason,
                "reported_by": lead.assigned_to,
                "created_by": "Admin"
            }
        )
        if not created:
            if main_reason_obj:
                loss_obj.main_reason = main_reason_obj
            loss_obj.detailed_reason = detailed_reason
            loss_obj.save()

        return {
            "status": "success",
            "message": f"Lead '{lead.full_name}' marked as LOST successfully!",
            "lead_id": lead.id,
            "stage": "Lost"
        }

    except Exception as e:
        raise APIException(str(e))
    
    
    
    
# ------------------------------edit lead admin-----------------------------

def edit_lead_admin(**data):
    """
    Admin Leads Page -> Edit Lead Modal Save API.
    Updates Lead Info and PaymentInfo records safely.
    """
    try:
        lead_id = data.get("lead_id") or data.get("id")
        if not lead_id:
            raise APIException("Lead ID is required")

        lead = Lead.objects.filter(id=lead_id).first()
        if not lead:
            raise APIException("Lead not found")

        # 1. Name Resolution
        full_name = data.get("full_name") or data.get("name")
        if not full_name:
            fn = data.get("first_name") or ""
            ln = data.get("last_name") or ""
            full_name = f"{fn} {ln}".strip()
        if full_name:
            lead.full_name = full_name
            if data.get("first_name"):
                lead.first_name = data.get("first_name")
            if data.get("last_name"):
                lead.last_name = data.get("last_name")

        # 2. Basic Info Updates
        if data.get("mobile_no"):
            lead.mobile_no = data.get("mobile_no")
        if "email" in data and data.get("email") is not None:
            lead.email = data.get("email")
        if "alt_mobile" in data and data.get("alt_mobile") is not None:
            lead.alternative_mobile = data.get("alt_mobile")
        if data.get("enquiry_date"):
            try:
                lead.enquiry_date = data.get("enquiry_date")
            except Exception:
                pass

        # 3. Foreign Key Resolutions (Supports both Names and IDs)
        assigned_val = data.get("assigned_to")
        if assigned_val:
            user_obj = User.objects.filter(Q(id=assigned_val if str(assigned_val).isdigit() else 0) | Q(username__icontains=assigned_val) | Q(first_name__icontains=assigned_val)).first()
            if user_obj:
                lead.assigned_to = user_obj

        plan_val = data.get("course_plan")
        if plan_val:
            plan_obj = CoursePlan.objects.filter(Q(id=plan_val if str(plan_val).isdigit() else 0) | Q(courseplan__icontains=plan_val)).first()
            if plan_obj:
                lead.course_plan = plan_obj

        course_val = data.get("course")
        if course_val:
            course_name_obj = CourseName.objects.filter(Q(id=course_val if str(course_val).isdigit() else 0) | Q(coursename__icontains=course_val)).first()
            if course_name_obj:
                lead.course_name = course_name_obj

        campaign_val = data.get("campaign")
        if campaign_val:
            camp_obj = CampaignName.objects.filter(Q(id=campaign_val if str(campaign_val).isdigit() else 0) | Q(name__icontains=campaign_val)).first()
            if camp_obj:
                lead.campaign = camp_obj

        stage_val = data.get("stage") or data.get("pipeline")
        if stage_val:
            stage_obj = PipelineStage.objects.filter(Q(id=stage_val if str(stage_val).isdigit() else 0) | Q(name__icontains=stage_val)).first()
            if stage_obj:
                lead.pipeline_stage = stage_obj

        lead.save()

        # 4. Payment Info Updates & Auto-sync with Pending Payments Page
        amount_paid = data.get("amount_paid")
        pending_amount = data.get("pending_amount")
        total_amount = data.get("total_amount")

        if amount_paid is not None or pending_amount is not None or total_amount is not None:
            ap_val = float(amount_paid) if amount_paid is not None else 0.0
            pa_val = float(pending_amount) if pending_amount is not None else 0.0

            payment_obj, created = PaymentInfo.objects.get_or_create(
                lead=lead,
                defaults={
                    'amount_paid': ap_val,
                    'pending_amount': pa_val,
                    'is_full_payment': (pa_val == 0)
                }
            )

            if amount_paid is not None:
                payment_obj.amount_paid = float(amount_paid)
            if pending_amount is not None:
                payment_obj.pending_amount = float(pending_amount)

            payment_obj.is_full_payment = (payment_obj.pending_amount == 0)
            payment_obj.save()

            final_ap = payment_obj.amount_paid
            final_pa = payment_obj.pending_amount
        else:
            existing_payment = PaymentInfo.objects.filter(lead=lead).first()
            final_ap = existing_payment.amount_paid if existing_payment else 0.0
            final_pa = existing_payment.pending_amount if existing_payment else 0.0

        return {
            "status": "success",
            "message": "Lead details updated successfully",
            "data": {
                "lead_id": lead.id,
                "amount_paid": final_ap,
                "pending_amount": final_pa
            }
        }

    except Exception as e:
        raise APIException(str(e))

    except Exception as e:
        raise APIException(str(e))