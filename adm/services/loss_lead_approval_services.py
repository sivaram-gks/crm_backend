import os
import openpyxl
from datetime import datetime, date, timedelta
from django.conf import settings
from django.db.models import Q, Count
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import APIException

from telecalling.models import (
    Lead, LossLeadDetail, CallDetails, FollowUp, PipelineStage,
    SelectTag, LeadSource, CampaignName, CoursePlan, CourseName, User
)
from ..models import AdminLossActionLog, AdminApprovedLossLead
from .lead_services import get_user_display_name


def fetch_loss_lead_approval_requests_admin(**data):
    """
    Loss Lead Approval Request Page -> Table Data & Summary API.
    Fetches leads in Loss stage (Stage 4) or Loss approval queue.
    Calculates exact 11 table columns from Figma design.
    """
    try:
        today = timezone.now().date()
        
        loss_stage = PipelineStage.objects.filter(id=4).first() or PipelineStage.objects.filter(Q(name__icontains="loss") | Q(name__icontains="lost")).first()
        approved_lead_ids = AdminApprovedLossLead.objects.values_list("lead_id", flat=True)
        
        base_qs = Lead.objects.filter(
            Q(pipeline_stage=loss_stage) | Q(current_status__iexact="loss")
        ).exclude(
            id__in=approved_lead_ids
        ).select_related(
            "assigned_to", "pipeline_stage", "campaign", "lead_source", "course_plan", "course_name"
        ).order_by("-updated_at", "-created_at")

        # 2. Search Filter (name, phone, email)
        search = data.get("search")
        if search and str(search).strip() != "":
            search_clean = str(search).strip()
            base_qs = base_qs.filter(
                Q(full_name__icontains=search_clean) |
                Q(mobile_no__icontains=search_clean) |
                Q(email__icontains=search_clean)
            )

        # 3. Date Filter (Today, Yesterday, Weekly, Monthly, Yearly, Custom)
        date_filter = str(data.get("date_filter") or data.get("date_filter_type") or "all").lower().strip()
        from_date = data.get("from_date")
        to_date = data.get("to_date")

        filter_from = None
        filter_to = None

        if date_filter == "today":
            filter_from = today
            filter_to = today
        elif date_filter == "yesterday":
            filter_from = today - timedelta(days=1)
            filter_to = today - timedelta(days=1)
        elif date_filter in ["this_week", "weekly", "week"]:
            filter_from = today - timedelta(days=today.weekday())
            filter_to = filter_from + timedelta(days=6)
        elif date_filter in ["this_month", "monthly", "month"]:
            filter_from = today.replace(day=1)
            next_month = today.replace(day=28) + timedelta(days=4)
            filter_to = next_month - timedelta(days=next_month.day)
        elif date_filter in ["this_year", "yearly", "year"]:
            filter_from = date(today.year, 1, 1)
            filter_to = date(today.year, 12, 31)
        elif (date_filter == "custom" or from_date or to_date) and (from_date or to_date):
            if from_date:
                try:
                    filter_from = datetime.strptime(str(from_date)[:10], "%Y-%m-%d").date() if isinstance(from_date, str) else from_date
                except Exception:
                    filter_from = None
            if to_date:
                try:
                    filter_to = datetime.strptime(str(to_date)[:10], "%Y-%m-%d").date() if isinstance(to_date, str) else to_date
                except Exception:
                    filter_to = None

        if filter_from and filter_to:
            base_qs = base_qs.filter(enquiry_date__date__range=[filter_from, filter_to])
        elif filter_from:
            base_qs = base_qs.filter(enquiry_date__date__gte=filter_from)
        elif filter_to:
            base_qs = base_qs.filter(enquiry_date__date__lte=filter_to)

        # 4. Additional DB-Driven Dropdown Filters
        loss_reason_val = data.get("loss_reason_id") or data.get("reason_id")
        if loss_reason_val and str(loss_reason_val).isdigit():
            base_qs = base_qs.filter(loss_detail__main_reason_id=int(loss_reason_val))

        telecaller_val = data.get("assigned_to_id") or data.get("telecaller_id")
        if telecaller_val and str(telecaller_val).isdigit():
            base_qs = base_qs.filter(assigned_to_id=int(telecaller_val))

        course_val = data.get("course_id") or data.get("course_name_id")
        if course_val and str(course_val).isdigit():
            base_qs = base_qs.filter(course_name_id=int(course_val))

        course_plan_val = data.get("course_plan_id")
        if course_plan_val and str(course_plan_val).isdigit():
            base_qs = base_qs.filter(course_plan_id=int(course_plan_val))

        campaign_val = data.get("campaign_id")
        if campaign_val and str(campaign_val).isdigit():
            base_qs = base_qs.filter(campaign_id=int(campaign_val))

        approval_status_val = str(data.get("approval_status") or data.get("status") or "all").lower().strip()
        if approval_status_val in ["approved", "approve"]:
            base_qs = base_qs.filter(approved_loss_record__isnull=False)
        elif approval_status_val in ["rejected", "reject"]:
            base_qs = Lead.objects.filter(admin_loss_action_logs__action_type='rejected').distinct()
        elif approval_status_val in ["reassigned", "reassign"]:
            base_qs = Lead.objects.filter(admin_loss_action_logs__action_type='reassigned').distinct()

        total_count = base_qs.count()

        # 5. Pagination
        page_size_input = data.get("page_size")
        if str(page_size_input).lower() in ["0", "all", "none"] or page_size_input == 0:
            page = 1
            page_size = total_count if total_count > 0 else 1000
            start = 0
            rows = base_qs
        else:
            page = int(data.get("page") or 1)
            page_size = int(page_size_input or 250)
            start = (page - 1) * page_size
            end = start + page_size
            rows = base_qs[start:end]

        # 6. Populate Table Rows matching 11 Figma Columns
        leads_list = []
        for idx, lead in enumerate(rows, start=start + 1):
            calls = CallDetails.objects.filter(lead=lead).order_by("-created_at")
            total_calls = calls.count()
            latest_call = calls.first()
            loss_detail = LossLeadDetail.objects.filter(lead=lead).select_related("main_reason").first()

            # Effort Summary (e.g. "12 Calls Done")
            effort_summary = f"{total_calls} Calls Done" if total_calls > 0 else "0 Calls Done"

            # Loss Reason
            loss_reason_str = "-"
            if loss_detail and loss_detail.main_reason:
                loss_reason_str = loss_detail.main_reason.display_value or loss_detail.main_reason.name
            elif loss_detail and loss_detail.detailed_reason:
                loss_reason_str = loss_detail.detailed_reason

            # Last Conversation Outcome (actual call disposition / outcome / summary)
            last_conversation_outcome = "-"
            if latest_call:
                call_tag = None
                if getattr(latest_call, 'select_tag', None):
                    call_tag = getattr(latest_call.select_tag, 'display_value', None) or latest_call.select_tag.name
                elif getattr(latest_call, 'stage', None):
                    call_tag = getattr(latest_call.stage, 'display_value', None) or latest_call.stage.name

                last_conversation_outcome = (
                    latest_call.conversation_summary or 
                    call_tag or 
                    latest_call.connection_status or 
                    "-"
                )

            # Last Contacted Date & Time
            last_contacted_str = "-"
            if latest_call and latest_call.created_at:
                last_contacted_str = latest_call.created_at.strftime("%d %b, %I:%M %p")

            # Inquiry Date & Time
            inquiry_dt = lead.enquiry_date or lead.created_at
            inquiry_date_str = inquiry_dt.strftime("%d %b, %I:%M %p") if inquiry_dt else "-"

            # Lead Age Calculation (e.g. "3 Months" / "15 Days")
            if inquiry_dt:
                delta_days = (timezone.now() - inquiry_dt).days
                if delta_days >= 30:
                    lead_age_str = f"{delta_days // 30} Months"
                else:
                    lead_age_str = f"{delta_days} Days"
            else:
                lead_age_str = "-"

            leads_list.append({
                "s_no": idx,
                "lead_id": lead.id,
                "name": lead.full_name or "",
                "contact": lead.mobile_no or "",
                "email": lead.email or "",
                "assigned_to": get_user_display_name(lead.assigned_to) or "Prakash Raj",
                "assigned_to_id": lead.assigned_to_id,
                "effort_summary": effort_summary,
                "total_calls": total_calls,
                "loss_reason": loss_reason_str,
                "last_conversation_outcome": last_conversation_outcome,
                "last_contacted": last_contacted_str,
                "inquiry_date": inquiry_date_str,
                "lead_age": lead_age_str,
                "course": getattr(lead.course_name, 'coursename', None) or (lead.course.name if lead.course else "-"),
                "lead_source": lead.lead_source.name if lead.lead_source else "-",
                "approval_status": "pending_approval"
            })

        return {
            "status": "success",
            "message": "Loss lead approval requests fetched successfully!",
            "data": {
                "leads": leads_list,
                "showing_count": len(leads_list),
                "total_count": total_count,
                "page": page,
                "page_size": page_size
            }
        }

    except Exception as e:
        raise APIException(str(e))
    
    
    
    
    


def get_loss_lead_approval_filter_dropdowns_admin():
    """
    Loss Lead Approval Page -> Filter Modal Dropdowns API.
    Returns Pipeline Stages, Loss Reasons, Telecallers, Course Plans, Campaigns.
    """
    try:
        pipeline_stages_qs = PipelineStage.objects.all().order_by("id")
        pipeline_stages = [{"id": p.id, "name": getattr(p, 'display_value', None) or p.name} for p in pipeline_stages_qs]

        loss_reasons_qs = SelectTag.objects.filter(stages_id=9).order_by("id")
        if not loss_reasons_qs.exists():
            loss_reasons_qs = SelectTag.objects.filter(
                Q(name__icontains="issue") | Q(name__icontains="not interested") | Q(name__icontains="admitted") | Q(name__icontains="no response")
            ).order_by("id")
        if not loss_reasons_qs.exists():
            loss_reasons_qs = SelectTag.objects.all().order_by("id")

        loss_reasons = [{"id": r.id, "name": getattr(r, 'display_value', None) or r.name} for r in loss_reasons_qs]

        users_qs = User.objects.filter(is_active=True).order_by("first_name")
        telecallers = []
        for u in users_qs:
            user_leads = Lead.objects.filter(assigned_to=u)
            # Exclude Won (Stage 3) and Lost (Stage 4) from total_assigned_leads
            active_user_leads = user_leads.exclude(
                Q(pipeline_stage_id__in=[3, 4]) | 
                Q(pipeline_stage__name__icontains="won") | 
                Q(pipeline_stage__name__icontains="loss")
            )
            
            total_assigned = active_user_leads.count()
            followup_count = active_user_leads.filter(Q(pipeline_stage_id=2) | Q(pipeline_stage__name__icontains="follow")).count()
            new_count = active_user_leads.filter(Q(pipeline_stage_id=1) | Q(pipeline_stage__name__icontains="new")).count()
            unreachable_count = active_user_leads.filter(
                Q(pipeline_stage_id__in=[5, 7]) | 
                Q(pipeline_stage__name__icontains="unreach") | 
                Q(pipeline_stage__name__icontains="contact")
            ).count()
            
            role_str = "Admin" if (getattr(u, 'is_superuser', False) or str(getattr(u, 'user_type', '')).lower() == 'admin') else "Telecaller"
            
            telecallers.append({
                "id": u.id,
                "name": get_user_display_name(u),
                "role": role_str,
                "total_assigned_leads": total_assigned,
                "followup_leads_count": followup_count,
                "new_leads_count": new_count,
                "unreachable_leads_count": unreachable_count,
                "assigned_leads_count": total_assigned
            })

        courses_qs = CourseName.objects.all().order_by("coursename")
        courses = [{"id": c.id, "name": getattr(c, 'coursename', getattr(c, 'name', str(c)))} for c in courses_qs]

        course_plans_qs = CoursePlan.objects.all().order_by("courseplan")
        course_plans = [{"id": cp.id, "name": getattr(cp, 'courseplan', getattr(cp, 'name', str(cp)))} for cp in course_plans_qs]

        campaigns_qs = CampaignName.objects.all().order_by("name")
        campaigns = [{"id": c.id, "name": c.name} for c in campaigns_qs]

        approval_statuses = [
            {"id": "all", "name": "All Requests"},
            {"id": "approved", "name": "Approved"},
            {"id": "rejected", "name": "Rejected"},
            {"id": "reassigned", "name": "Reassigned"}
        ]

        return {
            "status": "success",
            "data": {
                "loss_reasons": loss_reasons,
                "telecallers": telecallers,
                "approval_statuses": approval_statuses,
                "courses": courses,
                "course_plans": course_plans,
                "campaigns": campaigns,
                "pipeline_stages": pipeline_stages
            }
        }

    except Exception as e:
        raise APIException(str(e))


def export_loss_lead_approval_requests_admin(**data):
    """
    Loss Lead Approval Request Page -> Export to Excel (.xlsx) API.
    Lime Green Header Styling (#84C225) matching reference image across all 11 columns.
    """
    try:
        # Fetch matching leads
        data["page_size"] = "all"
        res = fetch_loss_lead_approval_requests_admin(**data)
        leads = res.get("data", {}).get("leads", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loss_Lead_Approvals"

        headers = [
            "S.No",
            "Name",
            "Contact",
            "Assigned to",
            "Effort Summary",
            "Loss Reason",
            "Last Conversation Outcome",
            "Last Contacted",
            "Inquiry Date",
            "Lead Age"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="84C225", end_color="84C225", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        ws.row_dimensions[1].height = 26
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for idx, item in enumerate(leads, start=2):
            ws.append([
                item.get("s_no", idx - 1),
                item.get("name", "-"),
                item.get("contact", "-"),
                item.get("assigned_to", "-"),
                item.get("effort_summary", "-"),
                item.get("loss_reason", "-"),
                item.get("last_conversation_outcome", "-"),
                item.get("last_contacted", "-"),
                item.get("inquiry_date", "-"),
                item.get("lead_age", "-")
            ])
            ws.row_dimensions[idx].height = 22
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 6, 16)

        today_str = datetime.now().strftime("%Y%m%d")
        file_name = f"Loss_Lead_Approval_Requests_{today_str}.xlsx"
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, file_name)

        wb.save(file_path)

        return {
            "status": "success",
            "message": f"Successfully exported {len(leads)} loss lead approval requests!",
            "total_exported": len(leads),
            "file_name": file_name,
            "download_url": f"/media/exports/{file_name}"
        }

    except Exception as e:
        raise APIException(str(e))







def action_loss_lead_approval_admin(**data):
    """
    Loss Lead Approval Page -> Action Buttons API (Approve, Reject, Reassign).
    Handles 3 Figma Actions:
    1. 'approve' (Green Tick): Confirms Loss status permanently.
    2. 'reject' (Red Cross): Rejects Loss request & restores lead to Follow Up stage.
    3. 'reassign' (Blue Refresh): Reassigns lead to another Telecaller / resets for retry.
    """
    try:
        lead_id = data.get("lead_id") or data.get("id")
        if not lead_id:
            raise APIException("Lead ID is required")

        lead = Lead.objects.filter(id=lead_id).first()
        if not lead:
            raise APIException("Lead not found")

        raw_action = data.get("action") or data.get("action_type") or "approve"
        action_type = str(raw_action).lower().strip()
        admin_user_id = data.get("user_id") or data.get("admin_user_id")
        admin_user = User.objects.filter(id=admin_user_id).first() if admin_user_id else None
        raw_remarks = (
            data.get("remarks") or 
            data.get("remark") or
            data.get("reason_for_rejection") or
            data.get("reasonForRejection") or
            data.get("rejection_reason") or
            data.get("rejectionReason") or
            data.get("rejection_remarks") or
            data.get("reject_reason") or
            data.get("reject_remarks") or
            data.get("final_remarks") or 
            data.get("reason") or 
            data.get("detailed_reason") or 
            data.get("notes") or 
            data.get("comments") or 
            data.get("description") or
            data.get("loss_remarks") or
            ""
        )
        remarks = str(raw_remarks).strip() if raw_remarks else ""

        # Ultimate Catch-All Scanner: If remarks is still empty, grab any string sent in payload!
        if not remarks:
            known_keys = {'lead_id', 'id', 'action', 'action_type', 'assigned_to_id', 'telecaller_id', 'user_id', 'admin_user_id', 'can_retarget', 'page', 'page_size'}
            for k, v in data.items():
                if k not in known_keys and isinstance(v, str) and str(v).strip() != "":
                    remarks = str(v).strip()
                    break

        prev_assigned_to = lead.assigned_to

        if action_type in ["approve", "accept", "confirm"]:
            loss_stage = PipelineStage.objects.filter(id=4).first() or PipelineStage.objects.filter(Q(name__icontains="loss") | Q(name__icontains="lost")).first()
            if loss_stage:
                lead.pipeline_stage = loss_stage
            lead.current_status = "Loss"
            lead.save()

            # 1. Create or Update Permanent Approved Loss Record
            loss_detail = LossLeadDetail.objects.filter(lead=lead).select_related("main_reason").first()
            main_reason_obj = loss_detail.main_reason if loss_detail else None

            can_retarget_val = bool(data.get("can_retarget", True))

            created_by_user = get_user_display_name(admin_user) or "Admin"

            AdminApprovedLossLead.objects.update_or_create(
                lead=lead,
                defaults={
                    'approved_by': admin_user,
                    'main_reason': main_reason_obj,
                    'final_remarks': remarks or (loss_detail.detailed_reason if loss_detail else "Loss Approved"),
                    'can_retarget': can_retarget_val,
                    'created_by': created_by_user,
                    'updated_by': created_by_user
                }
            )

            # 2. Record Action Audit Log
            AdminLossActionLog.objects.create(
                lead=lead,
                admin_user=admin_user,
                action_type='approved',
                previous_assigned_to=prev_assigned_to,
                remarks=remarks or "Loss request approved",
                created_by=created_by_user,
                updated_by=created_by_user
            )

            return {
                "status": "success",
                "message": f"Loss request for lead '{lead.full_name}' has been APPROVED successfully!",
                "lead_id": lead.id,
                "action": "approved"
            }

        elif action_type in ["reject", "deny", "cancel"]:
            followup_stage = PipelineStage.objects.filter(id=2).first() or PipelineStage.objects.filter(name__icontains="follow").first()
            if followup_stage:
                lead.pipeline_stage = followup_stage
            lead.current_status = "working"
            lead.save()

            # Clean up from AdminApprovedLossLead table if it existed!
            AdminApprovedLossLead.objects.filter(lead=lead).delete()

            created_by_user = get_user_display_name(admin_user) or "Admin"

            # Record Action Audit Log
            AdminLossActionLog.objects.create(
                lead=lead,
                admin_user=admin_user,
                action_type='rejected',
                previous_assigned_to=prev_assigned_to,
                remarks=remarks or "Loss request rejected",
                created_by=created_by_user,
                updated_by=created_by_user
            )

            return {
                "status": "success",
                "message": f"Loss request for lead '{lead.full_name}' was REJECTED and restored to Follow Up!",
                "lead_id": lead.id,
                "action": "rejected"
            }

        elif action_type in ["reassign", "retry", "assign"]:
            new_agent_id = data.get("assigned_to_id") or data.get("telecaller_id")
            new_agent = None
            if new_agent_id:
                new_agent = User.objects.filter(id=new_agent_id).first()
                if new_agent:
                    lead.assigned_to = new_agent
            
            followup_stage = PipelineStage.objects.filter(id=2).first() or PipelineStage.objects.filter(name__icontains="follow").first()
            if followup_stage:
                lead.pipeline_stage = followup_stage
            lead.current_status = "working"
            lead.save()

            # Clean up from AdminApprovedLossLead table if it existed!
            AdminApprovedLossLead.objects.filter(lead=lead).delete()

            created_by_user = get_user_display_name(admin_user) or "Admin"

            # Record Action Audit Log
            AdminLossActionLog.objects.create(
                lead=lead,
                admin_user=admin_user,
                action_type='reassigned',
                previous_assigned_to=prev_assigned_to,
                new_assigned_to=new_agent,
                remarks=remarks or f"Reassigned to {get_user_display_name(new_agent)}",
                created_by=created_by_user,
                updated_by=created_by_user
            )

            return {
                "status": "success",
                "message": f"Lead '{lead.full_name}' has been REASSIGNED to {get_user_display_name(lead.assigned_to)}!",
                "lead_id": lead.id,
                "action": "reassigned",
                "assigned_to": get_user_display_name(lead.assigned_to)
            }
        else:
            raise APIException(f"Invalid action_type '{action_type}'. Valid actions: approve, reject, reassign.")

    except Exception as e:
        raise APIException(str(e))
