import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, date
from rest_framework.exceptions import APIException

from telecalling.models import (
    Lead, PaymentInfo, PaymentHistory, FollowUp, CallDetails,
    PipelineStage, User, CourseName, CoursePlan, CourseTiming, PaymentStage
)

from datetime import datetime, date, timedelta

def get_pending_payment_filter_dropdowns_admin():
    """
    Pending Payments Page -> Filter Modal Dropdowns API.
    Returns dropdown options for Course Name, Course Plan, Course Time, Payment Stage, Pending Amount Ranges.
    """
    try:
        course_names = [
            {"id": c.id, "name": c.coursename}
            for c in CourseName.objects.filter(is_active=True).order_by("coursename")
            if c.coursename
        ]

        course_plans = [
            {"id": cp.id, "name": cp.courseplan}
            for cp in CoursePlan.objects.filter(is_active=True).order_by("courseplan")
            if cp.courseplan
        ]

        course_timings = [
            {"id": ct.id, "name": ct.coursetime}
            for ct in CourseTiming.objects.filter(is_active=True).order_by("coursetime")
            if ct.coursetime
        ]

        payment_stages = [
            {"id": "today_due", "name": "Today Due"},
            {"id": "active_due", "name": "Active Due"},
            {"id": "overdue", "name": "Overdue"}
        ]

        pending_amount_ranges = [
            {"id": "all", "name": "All"},
            {"id": "above_5k", "name": "Above ₹5,000"},
            {"id": "below_5k", "name": "Below ₹5,000"},
            {"id": "below_2k", "name": "Below ₹2,000"}
        ]

        return {
            "status": "success",
            "course_names": course_names,
            "course_plans": course_plans,
            "course_timings": course_timings,
            "payment_stages": payment_stages,
            "pending_amount_ranges": pending_amount_ranges
        }

    except Exception as e:
        raise APIException(str(e))


def fetch_all_pending_payments_admin(search=None, date_filter=None, from_date=None, to_date=None, date_filter_type=None, sort_by=None, pipeline_id=None, course_name_id=None, course_plan_id=None, course_timing_id=None, payment_stage_id=None, pending_amount_range=None, page=1, limit=1000):
    """
    Pending Payments Page -> Fetch All Pending Payments & Summary Cards API.
    Supports search, pipeline_id, Date Filters, Course Name, Course Plan, Course Time, Pending Amount Range.
    """
    try:
        date_filter = date_filter or date_filter_type
        today = timezone.now().date()
        if not limit or int(limit) <= 0:
            limit = 1000
        else:
            limit = int(limit)
        page = int(page) if page else 1

        # 1. Base Query: Only leads with pending payments > 0
        payments_qs = PaymentInfo.objects.select_related(
            'lead', 'lead__assigned_to', 'lead__pipeline_stage', 
            'lead__campaign', 'lead__course_plan', 'lead__course_name', 'lead__course', 'lead__course_timing'
        ).filter(pending_amount__gt=0)

        # 2. Search Filter (Name, Mobile, Email)
        if search:
            search = str(search).strip()
            payments_qs = payments_qs.filter(
                Q(lead__full_name__icontains=search) |
                Q(lead__mobile_no__icontains=search) |
                Q(lead__email__icontains=search)
            )

        # 3. Pipeline Filter
        if pipeline_id:
            payments_qs = payments_qs.filter(lead__pipeline_stage_id=pipeline_id)

        # 4. Course Filters
        if course_name_id:
            payments_qs = payments_qs.filter(
                Q(lead__course_name_id=course_name_id) | Q(lead__course__name_id=course_name_id)
            )

        if course_plan_id:
            payments_qs = payments_qs.filter(
                Q(lead__course_plan_id=course_plan_id) | Q(lead__course__plan_id=course_plan_id)
            )

        if course_timing_id:
            payments_qs = payments_qs.filter(
                Q(lead__course_timing_id=course_timing_id) | Q(lead__course__time_id=course_timing_id)
            )

        # 5. Pending Amount Range Filter
        if pending_amount_range == "below_5000":
            payments_qs = payments_qs.filter(pending_amount__lt=5000)
        elif pending_amount_range == "5000_10000":
            payments_qs = payments_qs.filter(pending_amount__gte=5000, pending_amount__lte=10000)
        elif pending_amount_range == "above_10000":
            payments_qs = payments_qs.filter(pending_amount__gt=10000)

        # 4. Resolve Due Date for all pending payments in memory to guarantee 100% accuracy
        all_payments = list(payments_qs.order_by('-id'))

        total_pending_amount = 0.0
        total_pending_leads = len(all_payments)

        due_today_amount = 0.0
        due_today_leads = 0

        overdue_amount = 0.0
        overdue_leads = 0

        processed_leads = []

        for idx, p in enumerate(all_payments, start=1):
            lead = p.lead

            # Fallback Due Date Resolution:
            # Order: 1. PaymentHistory.due_date -> 2. FollowUp.scheduled_at -> 3. Lead.created_at
            latest_ph = p.payment_histories.order_by('-id').first()
            p_due_date = latest_ph.due_date if (latest_ph and latest_ph.due_date) else None

            if not p_due_date:
                next_f = FollowUp.objects.filter(lead=lead, is_attended=False).order_by('scheduled_at').first()
                if next_f and next_f.scheduled_at:
                    p_due_date = next_f.scheduled_at.date()

            if not p_due_date and lead and lead.created_at:
                p_due_date = lead.created_at.date()

            # Status Determination & Summary Aggregations
            pending_amt = float(p.pending_amount or 0.0)
            total_pending_amount += pending_amt

            if p_due_date:
                if p_due_date < today:
                    status_text = "Overdue"
                    overdue_amount += pending_amt
                    overdue_leads += 1
                elif p_due_date == today:
                    status_text = "Due Today"
                    due_today_amount += pending_amt
                    due_today_leads += 1
                else:
                    status_text = "Active"
            else:
                status_text = "Active"

            # Date Filter Match Check
            filter_str = str(date_filter or "").lower().strip()
            filter_from = None
            filter_to = None

            if filter_str == "today":
                filter_from = today
                filter_to = today
            elif filter_str == "yesterday":
                filter_from = today - timedelta(days=1)
                filter_to = today - timedelta(days=1)
            elif filter_str in ["this_week", "weekly", "week"]:
                filter_from = today - timedelta(days=today.weekday())
                filter_to = filter_from + timedelta(days=6)
            elif filter_str in ["this_month", "monthly", "month"]:
                filter_from = today.replace(day=1)
                next_month = today.replace(day=28) + timedelta(days=4)
                filter_to = next_month - timedelta(days=next_month.day)
            elif filter_str in ["this_year", "yearly", "year"]:
                filter_from = date(today.year, 1, 1)
                filter_to = date(today.year, 12, 31)
            elif (filter_str == "custom" or from_date or to_date) and (from_date or to_date):
                if from_date:
                    try:
                        filter_from = datetime.strptime(str(from_date)[:10], "%Y-%m-%d").date() if isinstance(from_date, str) else from_date
                    except Exception:
                        filter_from = None
                if to_date:
                    try:
                        filter_to = datetime.strptime(str(to_date)[:10], "%Y-%m-%d").date() if isinstance(to_date, str) else to_date
                    except Exception:
                        filter_to = None

            if filter_str == "overdue":
                if not p_due_date or p_due_date >= today:
                    continue
            elif filter_str == "upcoming":
                if not p_due_date or p_due_date <= today:
                    continue
            elif filter_from or filter_to:
                if not p_due_date:
                    continue
                if filter_from and p_due_date < filter_from:
                    continue
                if filter_to and p_due_date > filter_to:
                    continue

            # Fetch Next Followup
            next_f_obj = FollowUp.objects.filter(lead=lead, is_attended=False).order_by('scheduled_at').first()
            next_followup_str = next_f_obj.scheduled_at.strftime("%d %b, %I:%M %p") if (next_f_obj and next_f_obj.scheduled_at) else None

            # Fetch Last Conversation
            last_call = CallDetails.objects.filter(lead=lead).order_by('-called_at').first()
            last_conv = last_call.conversation_summary if (last_call and last_call.conversation_summary) else (next_f_obj.notes if next_f_obj else "No conversation recorded yet")

            # Assigned User Name
            assigned_name = "-"
            if lead and lead.assigned_to:
                fname = (lead.assigned_to.first_name or "").strip()
                lname = (lead.assigned_to.last_name or "").strip()
                assigned_name = f"{fname} {lname}".strip() or lead.assigned_to.username or "-"

            # Formatting Enrolled / Joining Date
            joining_date_str = lead.created_at.strftime("%d %b, %I:%M %p") if (lead and lead.created_at) else "-"

            processed_leads.append({
                "s_no": idx,
                "id": lead.id if lead else p.id,
                "payment_id": p.id,
                "name": lead.full_name if lead else "-",
                "contact": lead.mobile_no if lead else "-",
                "email": lead.email if lead else "-",
                "assigned_to": assigned_name,
                "pipeline": lead.pipeline_stage.name.title() if (lead and lead.pipeline_stage) else "-",
                "campaign": lead.campaign.name if (lead and lead.campaign) else "-",
                "course_plan": getattr(lead.course_plan, 'courseplan', '-') if (lead and lead.course_plan) else "-",
                "course": getattr(lead.course_name, 'coursename', None) or (getattr(lead.course.name, 'coursename', '-') if (lead and lead.course and lead.course.name) else "-"),
                "joining_date": joining_date_str,
                "batch_timing": getattr(lead.course_timing, 'coursetime', '-') if (lead and hasattr(lead, 'course_timing') and lead.course_timing) else "Standard Batch",
                "amount_paid": float(p.amount_paid or 0.0),
                "pending_amount": pending_amt,
                "status": status_text,
                "due_date": p_due_date.strftime("%Y-%m-%d") if p_due_date else None,
                "next_followup": next_followup_str,
                "last_conversation": last_conv,
                "_raw_due_date": p_due_date
            })

        # Sorting
        if sort_by == "due_date_asc":
            processed_leads.sort(key=lambda x: x['_raw_due_date'] or date.max)
        elif sort_by == "due_date_desc":
            processed_leads.sort(key=lambda x: x['_raw_due_date'] or date.min, reverse=True)
        elif sort_by == "amount_desc":
            processed_leads.sort(key=lambda x: x['pending_amount'], reverse=True)
        elif sort_by == "amount_asc":
            processed_leads.sort(key=lambda x: x['pending_amount'])

        # Clean up temporary raw sorting key
        for item in processed_leads:
            item.pop('_raw_due_date', None)

        summary_cards = {
            "total_pending": {
                "amount": total_pending_amount,
                "total_amount": total_pending_amount,
                "value": total_pending_amount,
                "pending_amount": total_pending_amount,
                "leads_count": total_pending_leads,
                "count": total_pending_leads,
                "leads": total_pending_leads,
                "total_leads": total_pending_leads,
                "label": "Total Pending"
            },
            "due_today": {
                "amount": due_today_amount,
                "total_amount": due_today_amount,
                "value": due_today_amount,
                "pending_amount": due_today_amount,
                "leads_count": due_today_leads,
                "count": due_today_leads,
                "leads": due_today_leads,
                "total_leads": due_today_leads,
                "label": "Due Today"
            },
            "overdue_amount": {
                "amount": overdue_amount,
                "total_amount": overdue_amount,
                "value": overdue_amount,
                "pending_amount": overdue_amount,
                "leads_count": overdue_leads,
                "count": overdue_leads,
                "leads": overdue_leads,
                "total_leads": overdue_leads,
                "alert": "Need Immediate Action",
                "label": "Overdue Amount"
            }
        }

        # Pagination
        total_count = len(processed_leads)
        start = (page - 1) * limit
        end = start + limit
        paginated_leads = processed_leads[start:end]

        # Reset s_no for paginated items
        for i, lead_item in enumerate(paginated_leads, start=start + 1):
            lead_item['s_no'] = i

        return {
            "status": "success",
            # Exact number properties expected by React PendingPaymentStats.jsx
            "total_pending_amount": total_pending_amount,
            "total_leads": total_pending_leads,
            "today_due_amount": due_today_amount,
            "today_due_leads": due_today_leads,
            "due_today_amount": due_today_amount,
            "due_today_leads": due_today_leads,
            "overdue_amount": overdue_amount,
            "overdue_leads": overdue_leads,
            "summary_cards": summary_cards,
            "total_count": total_count,
            "page": page,
            "limit": limit,
            "leads": paginated_leads
        }

    except Exception as e:
        raise APIException(str(e))


def export_pending_payments_admin(
    search=None, date_filter=None, from_date=None, to_date=None, date_filter_type=None,
    sort_by=None, pipeline_id=None, course_name_id=None, course_plan_id=None,
    course_timing_id=None, payment_stage_id=None, pending_amount_range=None, **kwargs
):
    """
    Pending Payments Page -> Export Pending Payments Excel API.
    Features:
      - Header Fill: Lime Green (#84C225)
      - Header Font: Bold White (#FFFFFF)
      - Alignment: Center & Middle for all columns
      - Column Widths: Auto-calculated with +6 padding
      - Returns: Binary Excel File download response (.xlsx)
    """
    try:
        today = timezone.now().date()

        # Fetch pending payments data using existing service logic
        res = fetch_all_pending_payments_admin(
            search=search,
            date_filter=date_filter,
            from_date=from_date,
            to_date=to_date,
            date_filter_type=date_filter_type,
            sort_by=sort_by,
            pipeline_id=pipeline_id,
            course_name_id=course_name_id,
            course_plan_id=course_plan_id,
            course_timing_id=course_timing_id,
            payment_stage_id=payment_stage_id,
            page=1,
            limit=10000
        )
        leads = res.get("leads", [])

        # Create OpenPyXL Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pending Payments"
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        header_fill = PatternFill(start_color="84C225", end_color="84C225", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Calibri", size=10, color="000000")
        data_align = Alignment(horizontal="center", vertical="center")

        thin_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # 17 Columns Headers
        headers = [
            "S.No",
            "Student Name",
            "Contact Number",
            "Email Address",
            "Assigned Telecaller",
            "Pipeline Stage",
            "Campaign",
            "Course Plan",
            "Course Name",
            "Joining Date",
            "Batch & Timing",
            "Amount Paid (₹)",
            "Pending Amount (₹)",
            "Payment Status",
            "Due Date",
            "Next Followup",
            "Last Conversation Summary"
        ]

        ws.row_dimensions[1].height = 32
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Add Data Rows
        for row_idx, lead in enumerate(leads, start=2):
            row_data = [
                lead.get("s_no", row_idx - 1),
                lead.get("name", "-"),
                lead.get("contact", "-"),
                lead.get("email") or "-",
                lead.get("assigned_to", "-"),
                lead.get("pipeline", "-"),
                lead.get("campaign", "-"),
                lead.get("course_plan", "-"),
                lead.get("course", "-"),
                lead.get("joining_date", "-"),
                lead.get("batch_timing", "-"),
                lead.get("amount_paid", 0.0),
                lead.get("pending_amount", 0.0),
                lead.get("status", "-"),
                lead.get("due_date") or "-",
                lead.get("next_followup") or "-",
                lead.get("last_conversation", "-")
            ]

            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 25

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        # Auto-adjust Column Widths (+6 padding)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 6, 15)

        # Create HTTP Response for Excel Download
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Pending_Payments_Report_{today.strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response

    except Exception as e:
        raise APIException(str(e))
